import openpyxl

def split_excel_skip_hidden(file_path):
    """
    通过加载源文件、删除其他工作表并另存的方式，完整地拆分Excel文件。
    此版本会检测并自动跳过所有隐藏的工作表。

    参数:
    file_path (str): 要拆分的 .xlsx 文件的路径。
    """
    try:
        # 步骤 1: 加载工作簿一次，用于分析其所有工作表的状态
        print(f"正在分析文件: '{file_path}'...")
        source_workbook = openpyxl.load_workbook(file_path)
        
        visible_sheets_to_process = []
        hidden_sheets_skipped = []

        # 遍历所有工作表，将它们分类为“可见”或“隐藏”
        for sheet_name in source_workbook.sheetnames:
            sheet = source_workbook[sheet_name]
            if sheet.sheet_state == 'visible':
                visible_sheets_to_process.append(sheet_name)
            else:
                hidden_sheets_skipped.append(sheet_name)
        
        source_workbook.close() # 分析完毕，关闭文件以释放内存

        # 步骤 2: 向用户报告分析结果
        if not visible_sheets_to_process:
            print("错误：在文件中没有找到任何可见的工作表可供拆分。")
            return

        print(f"在文件中找到可见工作表: {', '.join(visible_sheets_to_process)}")
        if hidden_sheets_skipped:
            print(f"将跳过以下隐藏的工作表: {', '.join(hidden_sheets_skipped)}")
        print("-" * 30)

        # 步骤 3: 仅遍历并处理“可见”的工作表列表
        for sheet_to_keep in visible_sheets_to_process:
            print(f"正在处理可见工作表: '{sheet_to_keep}'...")
            
            # 为每个需要保存的可见工作表，重新加载一次原始文件
            wb_copy = openpyxl.load_workbook(file_path)
            
            # 从这个副本中删除所有其他工作表
            for sheet_name in list(wb_copy.sheetnames):
                if sheet_name != sheet_to_keep:
                    wb_copy.remove(wb_copy[sheet_name])
            
            # 现在，副本中只剩下一个我们已知的可见工作表
            # 定义输出文件名并保存
            output_file_name = f"{sheet_to_keep}.xlsx"
            wb_copy.save(output_file_name)
            wb_copy.close()

            print(f"✔ 已成功创建文件: '{output_file_name}'")

        print("-" * 30)
        print("所有可见工作表均已处理完毕。")

    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'。请确保文件路径正确。")
    except Exception as e:
        print(f"处理过程中发生错误：{e}")

# --- 使用示例 ---
if __name__ == "__main__":
    # 将 "example_table.xlsx" 替换为您要拆分的实际文件名
    input_excel_file = "../example_table.xlsx" 
    split_excel_skip_hidden(input_excel_file)